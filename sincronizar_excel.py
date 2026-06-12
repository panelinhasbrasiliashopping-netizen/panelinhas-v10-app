#!/usr/bin/env python3
"""
sincronizar_excel.py
Rode este script para sincronizar os dados do app com o Excel.
Uso: python sincronizar_excel.py
"""
import json, re, sys, time
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')
from pathlib import Path
from datetime import datetime

print("=" * 55)
print("  SINCRONIZAÇÃO — App Web → Excel  ")
print("=" * 55)

# ── 1. Dependências ──────────────────────────────────
try:
    import gspread
    from google.oauth2.service_account import Credentials
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError as e:
    print(f"\n❌ Biblioteca faltando: {e}")
    print("   Rode: python -m pip install gspread google-auth openpyxl")
    sys.exit(1)

# ── 2. Credenciais ───────────────────────────────────
secrets_path = Path(__file__).parent / ".streamlit" / "secrets.toml"
if not secrets_path.exists():
    print(f"\n❌ Arquivo não encontrado: {secrets_path}")
    sys.exit(1)

content = secrets_path.read_text(encoding="utf-8")

def extract(field):
    m = re.search(rf'{field}\s*=\s*"([^"]+)"', content)
    return m.group(1) if m else ""

key_match = re.search(r'private_key\s*=\s*"""(.*?)"""', content, re.DOTALL)
pk = key_match.group(1).replace("\\n", "\n") if key_match else ""

SPREADSHEET_ID = extract("spreadsheet_id")
info = {
    "type": "service_account",
    "project_id":   extract("project_id"),
    "private_key_id": extract("private_key_id"),
    "private_key":  pk,
    "client_email": extract("client_email"),
    "client_id":    extract("client_id"),
    "auth_uri":     extract("auth_uri"),
    "token_uri":    extract("token_uri"),
    "auth_provider_x509_cert_url": extract("auth_provider_x509_cert_url"),
    "client_x509_cert_url": extract("client_x509_cert_url"),
}

# ── 3. Conecta ao Google Sheets ──────────────────────
print("\n🔌 Conectando ao Google Sheets...")
try:
    scopes = ["https://www.googleapis.com/auth/spreadsheets",
              "https://www.googleapis.com/auth/drive"]
    creds  = Credentials.from_service_account_info(info, scopes=scopes)
    gc     = gspread.authorize(creds)
    sh     = gc.open_by_key(SPREADSHEET_ID)
    print("   ✅ Conectado!")
except Exception as e:
    print(f"   ❌ Erro: {e}")
    sys.exit(1)

# ── 4. Lê dados do Sheets ────────────────────────────
print("\n📥 Lendo dados do Google Sheets...")

def ler_aba(nome):
    try:
        ws   = sh.worksheet(nome)
        rows = ws.get_all_records()
        print(f"   ✅ {nome}: {len(rows)} registros")
        return rows
    except Exception as e:
        print(f"   ⚠️  {nome}: {e}")
        return []

movimentos   = ler_aba("MOVIMENTOS")
estoque      = ler_aba("ESTOQUE")
itens        = ler_aba("ITENS")
fornecedores = ler_aba("FORNECEDORES")

if not movimentos and not estoque:
    print("\n⚠️  Nenhum dado encontrado no Sheets.")
    print("   Faça lançamentos no app e sincronize primeiro.")
    sys.exit(0)

entradas = [m for m in movimentos if m.get("tipo") == "ENTRADA"]
saidas   = [m for m in movimentos if m.get("tipo") == "SAÍDA"]
ajustes  = [m for m in movimentos if m.get("tipo") in ("AJUSTE_POS","AJUSTE_NEG")]

print(f"\n   📊 Entradas: {len(entradas)} | Saídas: {len(saidas)} | Ajustes: {len(ajustes)}")

# ── 5. Abre o Excel ──────────────────────────────────
import glob, os, re
# Procura em várias pastas possíveis
pastas = [
    Path(__file__).parent,
    Path(__file__).parent.parent,
    Path.home() / "Downloads",
    Path.home() / "OneDrive" / "Desktop",
    Path.home() / "Desktop",
    Path.home() / "OneDrive",
]
excel_files = []
for pasta in pastas:
    found = glob.glob(str(pasta / "Panelinhas*.xlsx"))
    found += glob.glob(str(pasta / "Planilha*.xlsx"))
    found += glob.glob(str(pasta / "*Estoque*.xlsx"))
    excel_files += found

# Remove duplicatas
seen = set()
excel_files = [f for f in excel_files if not (f in seen or seen.add(f))]

# Filtra arquivos de backup, temporários ou pré-migração
excel_files = [
    f for f in excel_files 
    if "backup" not in f.lower() 
    and "pre_migracao" not in f.lower() 
    and not Path(f).name.startswith("~$")
]

# Prioriza arquivos no diretório local se houver algum válido localmente
diretorio_local = Path(__file__).parent.resolve()
arquivos_locais = [f for f in excel_files if Path(f).parent.resolve() == diretorio_local]
if arquivos_locais:
    excel_files = arquivos_locais

# Função de ordenação super robusta: prioriza arquivos com abas "csv&sheet="
def sort_key(filepath):
    name = Path(filepath).name.lower()
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        has_csv_sheet = "csv&sheet=MOVIMENTOS" in wb.sheetnames
        wb.close()
    except Exception:
        has_csv_sheet = False
    
    if has_csv_sheet:
        if "app_dados" in name:
            return (1, 0)
        if "planilha de estoque" in name:
            return (2, 2)
        match = re.search(r'v(\d+)', name)
        version = int(match.group(1)) if match else 1
        return (2, version)
    return (0, 0)

excel_files.sort(key=sort_key, reverse=True)

if not excel_files:
    caminho = input("\n📂 Digite o caminho completo do arquivo Excel: ").strip().strip('"')
else:
    print("\n📂 Arquivo Excel encontrado:")
    for i, f in enumerate(excel_files[:5], 1):
        print(f"   {i}. {f}")
    if len(excel_files) == 1:
        caminho = excel_files[0]
    else:
        escolha = input("   Escolha o número (ou Enter para o primeiro): ").strip()
        try:
            if escolha:
                idx = int(escolha) - 1
                if 0 <= idx < len(excel_files):
                    caminho = excel_files[idx]
                else:
                    print(f"   ⚠️ Opção inválida ({escolha}). Usando o primeiro arquivo da lista.")
                    caminho = excel_files[0]
            else:
                caminho = excel_files[0]
        except ValueError:
            print(f"   ⚠️ Opção inválida ({escolha}). Usando o primeiro arquivo da lista.")
            caminho = excel_files[0]

print(f"\n   📄 Usando: {Path(caminho).name}")
wb = None
while True:
    try:
        wb = openpyxl.load_workbook(caminho)
        break
    except PermissionError:
        print(f"\n❌ ERRO DE ACESSO: O arquivo '{Path(caminho).name}' está aberto ou bloqueado no Excel!")
        print("   Por favor, FECHE a planilha no Excel para continuar.")
        escolha = input("   Pressione ENTER para tentar novamente ou digite 'Q' para cancelar: ").strip().lower()
        if escolha == 'q':
            print("❌ Sincronização cancelada pelo usuário.")
            sys.exit(0)
    except Exception as e:
        print(f"   ❌ Erro ao abrir Excel: {e}")
        sys.exit(1)

# ── 6. Lógica de Atualização Condicional (Excel v7 vs Legado) ───
def data_fmt(ts):
    if not ts: return ""
    try:    return datetime.fromisoformat(str(ts))
    except: return str(ts)

is_v7 = "csv&sheet=MOVIMENTOS" in wb.sheetnames

if is_v7:
    print("\n⚡ Planilha ERP v7 detectada (usa abas 'csv&sheet=').")
    print("   Sincronizando dados brutos e preservando todas as abas de fórmulas...\n")
    
    def atualizar_csv_sheet(ws_name, data, header_row=2, data_start_row=3):
        if ws_name not in wb.sheetnames:
            return False
        
        ws = wb[ws_name]
        
        # 1. Limpa tudo da linha data_start_row em diante
        max_r = ws.max_row
        if max_r >= data_start_row:
            for r in range(data_start_row, max_r + 1):
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).value = None
        
        if not data:
            print(f"   ✅ {ws_name}: Limpo (sem dados)")
            return True
            
        # 2. Pega cabeçalhos da linha header_row
        headers = [ws.cell(header_row, col).value for col in range(1, ws.max_column + 1)]
        headers = [h for h in headers if h is not None]
        
        # 3. Escreve os dados
        adicionados = 0
        for idx, item in enumerate(data):
            r = data_start_row + idx
            for c_idx, h in enumerate(headers, 1):
                val = item.get(h, "")
                if val != "":
                    try:
                        if isinstance(val, str):
                            val_clean = val.strip()
                            if val_clean.replace(".", "", 1).isdigit():
                                if "." in val_clean:
                                    val = float(val_clean)
                                else:
                                    val = int(val_clean)
                    except ValueError:
                        pass
                if h == "timestamp" and val != "":
                    val = data_fmt(val)
                ws.cell(r, c_idx).value = val
            adicionados += 1
            
        print(f"   ✅ {ws_name}: {adicionados} registros atualizados")
        return True

    # Atualiza as 4 abas brutas do PowerQuery / Integração v7
    atualizar_csv_sheet("csv&sheet=MOVIMENTOS", movimentos, header_row=1, data_start_row=3)
    atualizar_csv_sheet("csv&sheet=ESTOQUE", estoque, header_row=2, data_start_row=3)
    atualizar_csv_sheet("csv&sheet=ITENS", itens, header_row=2, data_start_row=3)
    atualizar_csv_sheet("csv&sheet=FORNECEDORES", fornecedores, header_row=2, data_start_row=3)

else:
    print("\n🍂 Planilha antiga detectada. Usando fluxo de compatibilidade direta...")
    
    # ── 7. Lógica Legada: Atualiza MOVIMENTOS ─────────────────
    if "MOVIMENTOS" in wb.sheetnames and movimentos:
        print("\n✏️  Atualizando aba MOVIMENTOS...")
        ws_all = wb["MOVIMENTOS"]
        ids_all = set()
        for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row, values_only=True):
            if row[0]: ids_all.add(str(row[0]))
        proxima_all = 2
        for r in range(2, ws_all.max_row + 100):
            if not ws_all.cell(r,1).value:
                proxima_all = r; break
        add_all = 0
        for mov in movimentos:
            if str(mov.get("id","")) in ids_all: continue
            r = proxima_all + add_all
            try:
                ws_all.cell(r,1).value  = mov.get("id","")
                ws_all.cell(r,2).value  = data_fmt(mov.get("timestamp"))
                ws_all.cell(r,3).value  = mov.get("tipo","")
                ws_all.cell(r,4).value  = mov.get("id_item","")
                ws_all.cell(r,5).value  = mov.get("nome_item","")
                ws_all.cell(r,6).value  = float(mov.get("quantidade") or 0)
                ws_all.cell(r,7).value  = float(mov.get("valor_unit") or 0)
                ws_all.cell(r,8).value  = float(mov.get("total") or 0)
                ws_all.cell(r,9).value  = mov.get("fornecedor_id","")
                ws_all.cell(r,10).value = mov.get("motivo","")
                ws_all.cell(r,11).value = mov.get("operador","")
                ws_all.cell(r,12).value = mov.get("obs","")
                add_all += 1
            except: pass
        print(f"   ✅ {add_all} movimentos adicionados")

    # ── 8. Lógica Legada: Atualiza ENTRADAS ───────────────────
    if "ENTRADAS" in wb.sheetnames and entradas:
        print("\n✏️  Atualizando aba ENTRADAS...")
        ws_ent = wb["ENTRADAS"]
        
        ids_existentes = set()
        for row in ws_ent.iter_rows(min_row=2, max_row=ws_ent.max_row, values_only=True):
            if row[0]: ids_existentes.add(str(row[0]))
        
        proxima = 2
        for r in range(2, ws_ent.max_row + 100):
            v = ws_ent.cell(r, 1).value
            if not v:
                proxima = r
                break
        
        adicionadas = 0
        for mov in entradas:
            mov_id = str(mov.get("id",""))
            if mov_id in ids_existentes:
                continue
            
            r = proxima + adicionadas
            def sw(c, v): 
                try: ws_ent.cell(r, c).value = v
                except: pass
            sw(1,  mov.get("id",""))
            sw(2,  data_fmt(mov.get("timestamp")))
            sw(3,  mov.get("id_item",""))
            sw(4,  mov.get("nome_item",""))
            sw(5,  float(mov.get("quantidade") or 0))
            sw(6,  float(mov.get("valor_unit") or 0))
            sw(7,  float(mov.get("total") or 0))
            sw(8,  mov.get("fornecedor_id",""))
            sw(9,  mov.get("motivo","Compra"))
            sw(10, mov.get("operador",""))
            sw(11, mov.get("obs",""))
            adicionadas += 1
        
        print(f"   ✅ {adicionadas} entradas adicionadas")

    # ── 9. Lógica Legada: Atualiza SAIDAS ─────────────────────
    if "SAIDAS" in wb.sheetnames and saidas:
        print("\n✏️  Atualizando aba SAIDAS...")
        ws_sai = wb["SAIDAS"]
        
        ids_existentes = set()
        for row in ws_sai.iter_rows(min_row=2, max_row=ws_sai.max_row, values_only=True):
            if row[0]: ids_existentes.add(str(row[0]))
        
        proxima = 2
        for r in range(2, ws_sai.max_row + 100):
            v = ws_sai.cell(r, 1).value
            if not v:
                proxima = r
                break
        
        adicionadas = 0
        for mov in saidas:
            mov_id = str(mov.get("id",""))
            if mov_id in ids_existentes:
                continue
            
            r = proxima + adicionadas
            def sw(c, v):
                try: ws_sai.cell(r, c).value = v
                except: pass
            sw(1,  mov.get("id",""))
            sw(2,  data_fmt(mov.get("timestamp")))
            sw(3,  mov.get("id_item",""))
            sw(4,  mov.get("nome_item",""))
            sw(5,  float(mov.get("quantidade") or 0))
            sw(6,  mov.get("motivo",""))
            sw(7,  mov.get("operador",""))
            sw(8,  mov.get("obs",""))
            adicionadas += 1
        
        print(f"   ✅ {adicionadas} saídas adicionadas")

    # ── 10. Lógica Legada: Atualiza AJUSTES ───────────────────
    if "AJUSTES" in wb.sheetnames and ajustes:
        print("\n✏️  Atualizando aba AJUSTES...")
        ws_ajt = wb["AJUSTES"]

        merged = set()
        for rng in ws_ajt.merged_cells.ranges:
            for row in rng.cells:
                for r2, c2 in [row] if hasattr(row, '__iter__') else [(row.row, row.column)]:
                    merged.add((r2, c2))
        
        def safe_write(ws, row, col, val):
            if (row, col) not in merged:
                try: ws.cell(row, col).value = val
                except: pass

        proxima = 2
        for r in range(2, ws_ajt.max_row + 100):
            v = ws_ajt.cell(r, 1).value
            if not v and (r, 1) not in merged:
                proxima = r
                break

        adicionadas = 0
        for mov in ajustes:
            r = proxima + adicionadas
            safe_write(ws_ajt, r, 1, mov.get("id",""))
            safe_write(ws_ajt, r, 2, data_fmt(mov.get("timestamp")))
            safe_write(ws_ajt, r, 3, mov.get("id_item",""))
            safe_write(ws_ajt, r, 4, mov.get("nome_item",""))
            safe_write(ws_ajt, r, 5, "POSITIVO" if mov.get("tipo")=="AJUSTE_POS" else "NEGATIVO")
            safe_write(ws_ajt, r, 6, float(mov.get("quantidade") or 0))
            safe_write(ws_ajt, r, 7, mov.get("motivo",""))
            safe_write(ws_ajt, r, 8, mov.get("operador",""))
            safe_write(ws_ajt, r, 9, mov.get("obs",""))
            adicionadas += 1

        print(f"   ✅ {adicionadas} ajustes adicionados")

    # ── 11. Lógica Legada: Atualiza ESTOQUE_ATUAL ─────────────
    if "ESTOQUE_ATUAL" in wb.sheetnames and estoque:
        print("\n✏️  Atualizando aba ESTOQUE_ATUAL...")
        ws_est = wb["ESTOQUE_ATUAL"]
        ids_est = {}
        for r in range(2, ws_est.max_row + 1):
            v = ws_est.cell(r, 1).value
            if v: ids_est[str(v)] = r
        proxima_est = max(ids_est.values()) + 1 if ids_est else 2
        atualizados = 0
        for item in estoque:
            id_item = str(item.get("id_item",""))
            if not id_item: continue
            r = ids_est.get(id_item, proxima_est)
            if r == proxima_est: proxima_est += 1
            try:
                ws_est.cell(r,1).value = id_item
                ws_est.cell(r,2).value = item.get("nome","")
                ws_est.cell(r,3).value = item.get("categoria","")
                ws_est.cell(r,4).value = item.get("local","")
                ws_est.cell(r,5).value = item.get("unidade","")
                ws_est.cell(r,6).value = float(item.get("saldo_atual") or 0)
                ws_est.cell(r,7).value = item.get("ultima_atualizacao","")
                atualizados += 1
            except: pass
        print(f"   ✅ {atualizados} itens atualizados")

# ── 12. Salva Excel com backup organizado ────────────────
import shutil

pasta_principal  = Path(caminho).parent
pasta_backups    = pasta_principal / "backups"
nome_base        = Path(caminho).stem.split("_backup")[0]   # ex: Panelinhas_ERP_v7
agora            = datetime.now()
data_legivel     = agora.strftime("%d-%m-%Y_%Hh%M")        # ex: 26-05-2026_14h32

# Cria a pasta de backups se não existir
pasta_backups.mkdir(exist_ok=True)

# ── Move backups antigos espalhados na pasta principal ──
backups_antigos = list(pasta_principal.glob(f"{nome_base}_backup_*.xlsx"))
movidos = 0
for arq in backups_antigos:
    destino = pasta_backups / arq.name
    if not destino.exists():
        shutil.move(str(arq), str(destino))
        movidos += 1
    else:
        arq.unlink()   # duplicado — remove silenciosamente
        movidos += 1

if movidos:
    print(f"\n🗂️  {movidos} backup(s) antigo(s) movido(s) → pasta backups/")

# ── Cria o novo backup na pasta organizada ──────────────
nome_backup  = f"{nome_base}_backup_{data_legivel}.xlsx"
caminho_bkp  = pasta_backups / nome_backup
shutil.copy(caminho, str(caminho_bkp))
print(f"\n💾 Backup salvo em: backups/{nome_backup}")

# ── Salva a planilha principal ──────────────────────────
try:
    wb.calculation.calcMode = 'auto'
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass

while True:
    try:
        wb.save(caminho)
        print(f"✅ Planilha principal salva via openpyxl: {Path(caminho).name}")
        break
    except PermissionError:
        print(f"\n❌ ERRO DE SALVAMENTO: O arquivo '{Path(caminho).name}' foi aberto no Excel durante a sincronização!")
        print("   Por favor, FECHE a planilha no Excel para permitir que os dados sejam salvos.")
        escolha = input("   Pressione ENTER para tentar novamente ou digite 'Q' para descartar as alterações: ").strip().lower()
        if escolha == 'q':
            print("❌ Sincronização cancelada. Os dados NÃO foram salvos na planilha.")
            sys.exit(0)

# ── Força recálculo completo das fórmulas usando Excel (Windows) ──
excel = None
excel_wb = None
try:
    import win32com.client as win32
    import os
    print("\n🖥️  Iniciando Excel em segundo plano para recalcular fórmulas...")
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False
    excel.DisplayAlerts = False
    
    caminho_abs = os.path.abspath(caminho)
    excel_wb = excel.Workbooks.Open(caminho_abs)
    excel.CalculateFull()
    excel_wb.Save()
    print("   ✅ Fórmulas recalculadas e salvas com sucesso no Excel!")
except Exception as e:
    print(f"   ⚠️ Não foi possível recalcular automaticamente via Excel: {e}")
    print("      Sem problemas! Ao abrir o arquivo no Excel, ele deve recalcular automaticamente.")
finally:
    if excel_wb:
        try:
            excel_wb.Close(SaveChanges=True)
        except:
            pass
    if excel:
        try:
            excel.Quit()
        except:
            pass

print("\n" + "=" * 55)
print("  ✅ SINCRONIZAÇÃO E CÁLCULO CONCLUÍDOS!")
print("=" * 55)
print(f"\n  Entradas: {len(entradas)} | Saídas: {len(saidas)} | Ajustes: {len(ajustes)}")
print(f"  Backup:   backups/{nome_backup}")
print("  Abra o Excel para ver os dados atualizados.\n")
